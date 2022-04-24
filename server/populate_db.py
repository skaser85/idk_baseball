from openpyxl import load_workbook
from sqlite3 import connect, Cursor, Connection

def get_data(c: Cursor, sql: str, values: tuple = (), fetchone: bool = True):
    if fetchone:
        return c.execute(sql, values).fetchone()
    else:
        return c.execute(sql, values).fetchall()

def insert(cn: Connection, c: Cursor, sql: str):
    c.execute(sql)
    cn.commit()

def process_player_name(name: str) -> tuple[str]:
    # 'Enyel De Los Santos62'
    first = name.split(' ')[0]
    last = ' '.join(name.split(' ')[1:])
    no = ''.join([n for n in last if n.isdigit()])
    last = last[:len(last)-len(no)]
    last = last.replace('\\', '')
    return first, last, no

def get_team_id(c: Cursor, team_name: str) -> int:    
    city = team_name.split(' ')[0]
    name = ' '.join(team_name.split(' ')[1:])
    team_id = get_data(c,'SELECT * FROM Team WHERE City=? AND Name=?', (city, name))

    if team_id is None:
        city = ' '.join(team_name.split(' ')[0:2])
        name = ' '.join(team_name.split(' ')[2:])
        team_id = get_data(c,'SELECT * FROM Team WHERE City=? AND Name=?', (city, name))
        if team_id is None:
            raise ValueError(f'Could not find a team ID for {team_name}')
    
    return team_id[0]

wb = load_workbook(filename='players.xlsx')

conn = connect('idk_db.db')
cursor = conn.cursor()

player_fields = ("First Name", "Last Name", "Current Team", "Position", "Bats", "Throws", "Age", "Weight - Lbs", "Birth Place", "Jersey No.", "Height - Feet", "Height - Inches")

for sheetname in wb.sheetnames:
    team_name = sheetname
    team_id = get_team_id(cursor, team_name)

    ws = wb[sheetname]

    for i, row in enumerate(ws.iter_rows(),3):
        pic_url = ws.cell(row=i+1,column=1).value
        player_name = ws.cell(row=i+1,column=2).value
        pos = ws.cell(row=i+1,column=3).value
        bats = ws.cell(row=i+1,column=4).value
        throws = ws.cell(row=i+1,column=5).value
        age = ws.cell(row=i+1,column=6).value
        height = ws.cell(row=i+1,column=7).value
        weight = ws.cell(row=i+1,column=8).value
        birth_place = ws.cell(row=i+1,column=9).value
        if not pic_url or not pic_url.startswith('https:'):
            continue
        first_name, last_name, jersey_no = process_player_name(player_name)
        weight = weight.split(' ')[0]
        height_ft, height_in = height.split(' ')
        height_ft = height_ft[:-1]
        height_in = height_in[:-1]
        values = []
        values.append(f'"{first_name}"')
        values.append(f'"{last_name}"')
        values.append(f'"{team_id}"')
        values.append(f'"{pos}"')
        values.append(f'"{bats}"')
        values.append(f'"{throws}"')
        values.append(f'"{age}"')
        values.append(f'"{weight}"')
        values.append(f'"{birth_place}"')
        values.append(f'"{jersey_no}"')
        values.append(f'"{height_ft}"')
        values.append(f'"{height_in}"')
        sql = f'INSERT INTO Player {player_fields} VALUES ({", ".join(values)});'
        insert(conn, cursor, sql)
